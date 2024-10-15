from django.shortcuts import render
from .forms import PessoaRecadastramentoForm, ProcessoForm, InscricaoForm, _PessoaRecadastramentoForm
from .models import PessoaRecadastramento, Processo, Inscricao
from django.views.decorators.csrf import csrf_exempt
import json
from autenticacao.functions import validate_cpf
from django.http import JsonResponse
from django.core.exceptions import ValidationError
from django.core import serializers
from django.contrib.auth.decorators import login_required
from autenticacao.models import Pessoa
from django.http import HttpResponse
from .functions import is_servidor

@login_required
def index(request):
    if not is_servidor(request.pessoa):
        return HttpResponse('Acesso negado', status=403)
    context = {
        'titulo': 'Sala do Empreendedor - Serviços Internos',
        'form_pessoa_recadastramento': PessoaRecadastramentoForm(),
        'form_processo': ProcessoForm(),
        'form_inscricao': InscricaoForm(),
    }
    return render(request, 'recadastramento/index.html', context)

from django.db import IntegrityError
@login_required
def atualziacao_cadastral(request):
    pessoa = Pessoa.objects.get(user = request.user)
    error = json.dumps({})
    messages_ = []
    try:        
        pessoaR = PessoaRecadastramento.objects.get(cpf=pessoa.cpf)
    except:
        pessoaR= pessoa

    if request.method == 'POST':
        try:
            instante =  PessoaRecadastramento.objects.get(cpf=pessoa.cpf)
            form = _PessoaRecadastramentoForm(request.POST, instance=instante)
        except:
            form = _PessoaRecadastramentoForm(request.POST)

        if form.is_valid():
            pessoaR = form.save()
            pessoaR.user_register = request.user
            pessoaR.save()
            if 'n_inscricao' in request.POST:
                # print(request.POST['n_inscricao'])
                # print(request.POST.getlist('n_inscricao'))
                for i in request.POST.getlist('n_inscricao'):
                    try:
                        inscricao = Inscricao(
                            pessoa_recadastramento=pessoaR,
                            numero_inscricao=i
                        )
                        inscricao.full_clean()
                        inscricao.save()     
                        inscricao.user_register = request.user
                        inscricao.save()                       
                    except Exception as e:
                        error_message = str(e)
                        if 'numero_inscricao' in error_message:
                            if i == '':
                                messages_.append({
                                'msg': 'Inscrição inválida.',
                                'status': 'error'
                            })
                            else:
                                messages_.append({
                                    'msg': f'Erro ao cadastrar a inscrição {i}, pois essa inscrição já existe.',
                                    'status': 'error'
                                })
                        else:                           
                        
                            messages_.append({
                                'msg': 'Ocorreu um erro inesperado ao cadastrar a inscrição.',
                                'status': 'error'
                            })
            messages_.append({'msg':'Seu cadastro de contribuinte foi atualizado com sucesso!', 'status': 'success'})
            
        else:            
            messages_.append({'msg':'Erro ao atualizar cadastro. Verifique os campos.', 'status': 'error'})
            
    context = {
        'titulo': 'Sala do Empreendedor - Serviços Internos',        
        'pessoa': pessoaR,
        'error': error,
        'messages_': messages_,
        'servidor': is_servidor(request.pessoa)
    }
    return render(request, 'recadastramento/tela_contribuinte.html', context)

@login_required
def checkCPF(request):    
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        cpf = data.get('cpf')
        try:
            cpf = validate_cpf(cpf)
        except:
            response_data = {'exists': True, 'message': 'CPF inválido.'}
            return JsonResponse(response_data)
        try:
            pessoa = PessoaRecadastramento.objects.get(cpf=cpf)
            pessoa_json = serializers.serialize('json', [pessoa])
            response_data = {'exists': True, 'message': 'CPF já existe. Confira os dados para atualizar.', 'pessoa': pessoa_json}
        except Exception as e:
                print(e)
                response_data = {'exists': False}

        return JsonResponse(response_data)
    return JsonResponse({})

@login_required
def checkCPF2(request):    
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        cpf = data.get('cpf')
        try:
            cpf = validate_cpf(cpf)
        except:
            response_data = {'exists': False, 'message': 'CPF inválido.'}
            return JsonResponse(response_data)
        try:
            pessoa = PessoaRecadastramento.objects.get(cpf=cpf)
            pessoa_json = serializers.serialize('json', [pessoa])
            response_data = {'exists': True, 'message': 'Contribuinte localizado <i class="fa-solid fa-circle-check"></i>', 'pessoa': pessoa_json}
        except:
            response_data = {'exists': False, 'message': 'Contribuinte não localizado <i class="fa-solid fa-circle-xmark"></i>'}

        return JsonResponse(response_data)
    return JsonResponse({})

# def cadastrar_pessoa(request):
#     if request.method == 'POST':
#         form = PessoaRecadastramentoForm(request.POST)
#         if form.is_valid():
#             form.save()
#     return index(request)

@login_required
@csrf_exempt
def cadastrar_pessoa(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(data)
            try:
                cpf = data.get('cpf').replace('-', '').replace('.', '')
                pessoa = PessoaRecadastramento.objects.get(cpf=cpf)
                pessoa.cpf=cpf
                pessoa.cnpj=data.get('cnpj')
                pessoa.nome_do_contribuinte=data['nome_do_contribuinte']
                pessoa.celular=data.get('celular')
                pessoa.cep=data.get('cep')
                pessoa.rua=data.get('rua')
                pessoa.numero=data.get('numero')
                pessoa.complemento=data.get('complemento')
                pessoa.bairro=data.get('bairro')
                pessoa.cidade=data.get('cidade')
                pessoa.estado=data.get('estado')
                pessoa.email=data.get('email')
                pessoa.user_register = request.user
                pessoa.cadastro_interno = True

                pessoa.email_procurador = data.get('email_procurador')
                pessoa.telefone_procurador = data.get('telefone_procurador')
                pessoa.cpf_procurador = data.get('cpf_procurador')
                pessoa.nome_procurador = data.get('nome_procurador')

                pessoa.responsavel_tributario = data.get('responsavel_tributario')
                pessoa.cpf_responsavel = data.get('cpf_responsavel')
                pessoa.email_responsavel = data.get('email_responsavel')
                pessoa.telefone_responsavel = data.get('telefone_responsavel')
                pessoa.cep_responsavel = data.get('cep_responsavel')
                pessoa.rua_responsavel = data.get('rua_responsavel')
                pessoa.numero_responsavel = data.get('numero_responsavel')
                pessoa.complemento_responsavel = data.get('complemento_responsavel')
                pessoa.bairro_responsavel = data.get('bairro_responsavel')
                pessoa.cidade_responsavel = data.get('cidade_responsavel')
                pessoa.estado_responsavel = data.get('estado_responsavel')
                
                message = 'Contribuinte atualizado com sucesso!'
            except Exception as e:
                print(e)
                pessoa = PessoaRecadastramento(
                    cpf=data.get('cpf').replace('-', '').replace('.', ''),
                    responsavel_tributario=data['responsavel_tributario'],
                    cnpj=data.get('cnpj'),
                    nome_do_contribuinte=data['nome_do_contribuinte'],
                    celular=data.get('celular'),
                    cep=data.get('cep'),
                    rua=data.get('rua'),
                    numero=data.get('numero'),
                    complemento=data.get('complemento'),
                    bairro=data.get('bairro'),
                    cidade=data.get('cidade'),
                    estado=data.get('estado'),
                    email=data.get('email'),
                    user_register = request.user,
                    cadastro_interno = True,

                    email_procurador = data.get('email_procurador'),
                    telefone_procurador = data.get('telefone_procurador'),
                    cpf_procurador = data.get('cpf_procurador'),
                    nome_procurador = data.get('nome_procurador'),

                    cpf_responsavel = data.get('cpf_responsavel'),
                    email_responsavel = data.get('email_responsavel'),
                    telefone_responsavel = data.get('telefone_responsavel')

                )
                message = 'Cadastro realizado com sucesso!'
                
            pessoa.full_clean()  # Valida os campos antes de salvar
            pessoa.save()                                
            return JsonResponse({'message': message}, status=201)
        except ValidationError as e:
            errors = e.message_dict
            return JsonResponse({'errors': errors}, status=400)
        except Exception as e:
            print(e)
            errors = {}
            # errors = e.message_dict
            return JsonResponse({'error': errors}, status=500)
    else:
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    

@login_required
@csrf_exempt
def cadastrar_processo(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cpf = data.get('cpf_processo').replace('-', '').replace('.', '')
            processo = Processo(
                requerente=PessoaRecadastramento.objects.get(cpf=cpf),
                requerimento=data['requerimento'],
                ano=data.get('ano'),
                localizacao=data['localizacao'],                
            )            
            processo.user_register = request.user            
            processo.full_clean()  # Valida os campos antes de salvar
            processo.save()            
            return JsonResponse({'message': 'Cadastro realizado com sucesso!'}, status=201)
        except ValidationError as e:
            print(e)
            errors = e.message_dict
            return JsonResponse({'errors': errors}, status=400)
        except Exception as e:
            print(e)            
            errors = {
                'cpf_processo': 'Erro ao cadastrar processo. Verifique se o contribuinte já foi cadastrado.'
            
            }
            return JsonResponse({'error': errors}, status=500)
    else:
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
@login_required    
@csrf_exempt
def cadastrar_inscricao(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cpf = data.get('cpf_inscricao').replace('-', '').replace('.', '')
            inscricao = Inscricao(
                pessoa_recadastramento=PessoaRecadastramento.objects.get(cpf=cpf),
                numero_inscricao=data['numero_inscricao'],                
            )         
            inscricao.full_clean()  # Valida os campos antes de salvar
            inscricao.save()
            inscricao.user_register = request.user
            inscricao.cadastro_interno = True
            inscricao.save()
            return JsonResponse({'message': 'Cadastro realizado com sucesso!'}, status=201)
        except ValidationError as e:
            errors = e.message_dict
            return JsonResponse({'errors': errors}, status=400)
        except Exception as e:
            print(e)            
            errors = {
                'cpf_inscricao': 'Erro ao cadastrar processo. Verifique se o contribuinte já foi cadastrado.'
            
            }
            return JsonResponse({'error': errors}, status=500)
    else:
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
@login_required
def cadastar_usuarios_no_recadastramento(request):
    pessoas = Pessoa.objects.all()
    ja_cadastrados = []
    for pessoa in pessoas:
        try:
            pessoa_recadastramento = PessoaRecadastramento(
                cpf=pessoa.cpf,
                nome_do_contribuinte=pessoa.nome,
                celular=pessoa.telefone,
                cep=pessoa.cep,
                rua=pessoa.endereco,
                numero=pessoa.numero,
                complemento=pessoa.complemento,
                bairro=pessoa.bairro,            
                email=pessoa.email
            )
            pessoa_recadastramento.full_clean()
            pessoa_recadastramento.save()
        except:
            ja_cadastrados.append(str(pessoa.nome)+';'+str(pessoa.cpf)+';')
    
    
    file_path = '/home/eduardo/Documentos/Github/desenvolve_nf (cópia)/financas_recadastramento/cadastrados.txt'
    
    with open(file_path, 'w') as file:
        for cadastrado in ja_cadastrados:
            file.write(cadastrado)
            file.write("\n")
        file.write("####")
        file.write("\n")
        file.write(f"total_cadastrados: {pessoas.count() - len(ja_cadastrados)}; total_que_já_estavam: {len(ja_cadastrados)};")
    
    with open(file_path, 'r') as file:
        response = HttpResponse(file.read(), content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="cadastrados.txt"'
    
    return response

def listar_contribuintes(request):
    contribuintes = PessoaRecadastramento.objects.all()

    context={
        'contribuintes': contribuintes
    }

    return render(request, 'recadastramento/listar_contribuintes.html', context)

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

@login_required
def exportar_cadastro_to_excel(request):
    if not is_servidor(request.pessoa):
        return HttpResponse('Acesso negado', status=403)

    contribuintes = PessoaRecadastramento.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contribuintes'
    ws.append(['CPF do Contribuinte', 'CNPJ do Contribuinte', 'Nome do Contribuinte', 'Celular', 
               'CEP', 'Rua', 'Número', 'Complemento', 'Bairro', 'Cidade', 'Estado', 'E-mail', 'Inscrições',
                'CPF do Responsável Tributário', 'Nome do responsável Tributário', 'Celular do responsável',
                'Email do responsável', 'CEP do responsável', 'Rua responsável', 'Número do responsável', 'Complemento do responsável', 'Bairro do responsável', 'Cidade do responsável', 'Estado do responsável','CPF do procurador', 'Nome do procurador', 'Celular do procurador',
                'Email do procurador'])

    for contribuinte in contribuintes:
        inscricoes = Inscricao.objects.filter(pessoa_recadastramento=contribuinte)

        ws.append([
            contribuinte.cpf,
            contribuinte.cnpj,
            contribuinte.nome_do_contribuinte,
            contribuinte.celular,
            contribuinte.cep,
            contribuinte.rua,
            contribuinte.numero,
            contribuinte.complemento,
            contribuinte.bairro,
            contribuinte.cidade,
            contribuinte.estado,
            contribuinte.email,
            ', '.join([inscricao.numero_inscricao for inscricao in inscricoes]),
            contribuinte.cpf_responsavel,
            contribuinte.responsavel_tributario,
            contribuinte.telefone_responsavel,
            contribuinte.email_responsavel,
            contribuinte.cep_responsavel,
            contribuinte.rua_responsavel,
            contribuinte.numero_responsavel,
            contribuinte.complemento_responsavel,
            contribuinte.bairro_responsavel,
            contribuinte.cidade_responsavel,
            contribuinte.estado_responsavel,
            contribuinte.cpf_procurador,
            contribuinte.nome_procurador,
            contribuinte.telefone_procurador,
            contribuinte.email_procurador,           
        ])

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tab = Table(displayName='Contribuintes', ref=ws.dimensions)
    # style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    # tab.tableStyleInfo = style    
    ws.add_table(tab)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=contribuintes.xlsx'
    wb.save(response)

    return response