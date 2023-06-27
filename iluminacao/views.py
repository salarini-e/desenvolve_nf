from django.shortcuts import render, redirect
from .forms import *
from autenticacao.models import Pessoa
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.apps import apps
from .models import *
from settings.decorators import group_required
from django.contrib.auth.models import Group
from datetime import datetime
from django.db.models import Count, Sum

from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime

STATUS_CHOICES=(
        ('0','Novo'),
        ('1','Aguardando'),
        ('2','Em execução'),
        ('f','Finalizado')
    )
PRIORIDADE_CHOICES=(
        ('0','Normal'),
        ('1','Moderada'),
        ('2','Urgente'),
    )

@group_required('os_acesso')
def index(request):
    return render(request, 'os_index.html')

from django.db import connection

@login_required
@group_required('os_acesso')
def os_painel(request):    
    # meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    # bairros = OrdemDeServico.objects.values_list('bairro', flat=True).distinct()
    # data = []  

    # for bairro in bairros:
    #     total=OrdemDeServico.objects.filter(bairro=bairro, status__in=['0', '1', '2'], dt_solicitacao__year='2023').count()
    #     os_por_mes = OrdemDeServico.objects.filter(bairro=bairro, status__in=['0', '1', '2'], dt_solicitacao__year='2023').annotate(
    #         jan=Count('id', filter=models.Q(dt_solicitacao__month=1)),
    #         fev=Count('id', filter=models.Q(dt_solicitacao__month=2)),
    #         mar=Count('id', filter=models.Q(dt_solicitacao__month=3)),
    #         abr=Count('id', filter=models.Q(dt_solicitacao__month=4)),
    #         mai=Count('id', filter=models.Q(dt_solicitacao__month=5)),
    #         jun=Count('id', filter=models.Q(dt_solicitacao__month=6)),
    #         jul=Count('id', filter=models.Q(dt_solicitacao__month=7)),
    #         ago=Count('id', filter=models.Q(dt_solicitacao__month=8)),
    #         set=Count('id', filter=models.Q(dt_solicitacao__month=9)),
    #         out=Count('id', filter=models.Q(dt_solicitacao__month=10)),
    #         nov=Count('id', filter=models.Q(dt_solicitacao__month=11)),
    #         dez=Count('id', filter=models.Q(dt_solicitacao__month=12)),
    #     ).values('jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez')

    #     dt={'bairro': bairro, 'mes': os_por_mes, 'total': total}
    #     if not any(item['bairro'] == bairro for item in data):
    #         data.append(dt)
    # bairros = (
    #     OrdemDeServico.objects
    #     .filter(status__in=['0', '1', '2'], dt_solicitacao__year='2023')
    #     .values('bairro')
    #     .annotate(
    #         total=Count('id'),
    #         jan=Sum(models.Case(models.When(dt_solicitacao__month=1, then=1), default=0, output_field=models.IntegerField())),
    #         fev=Sum(models.Case(models.When(dt_solicitacao__month=2, then=1), default=0, output_field=models.IntegerField())),
    #         mar=Sum(models.Case(models.When(dt_solicitacao__month=3, then=1), default=0, output_field=models.IntegerField())),
    #         abr=Sum(models.Case(models.When(dt_solicitacao__month=4, then=1), default=0, output_field=models.IntegerField())),
    #         mai=Sum(models.Case(models.When(dt_solicitacao__month=5, then=1), default=0, output_field=models.IntegerField())),
    #         jun=Sum(models.Case(models.When(dt_solicitacao__month=6, then=1), default=0, output_field=models.IntegerField())),
    #         jul=Sum(models.Case(models.When(dt_solicitacao__month=7, then=1), default=0, output_field=models.IntegerField())),
    #         ago=Sum(models.Case(models.When(dt_solicitacao__month=8, then=1), default=0, output_field=models.IntegerField())),
    #         set=Sum(models.Case(models.When(dt_solicitacao__month=9, then=1), default=0, output_field=models.IntegerField())),
    #         out=Sum(models.Case(models.When(dt_solicitacao__month=10, then=1), default=0, output_field=models.IntegerField())),
    #         nov=Sum(models.Case(models.When(dt_solicitacao__month=11, then=1), default=0, output_field=models.IntegerField())),
    #         dez=Sum(models.Case(models.When(dt_solicitacao__month=12, then=1), default=0, output_field=models.IntegerField()))
    #     )
    # )

    # data = []
    # for bairro in bairros:
    #     bairro_data = {
    #         'bairro': bairro['bairro'],
    #         'mes': {
    #             'jan': bairro['jan'],
    #             'fev': bairro['fev'],
    #             'mar': bairro['mar'],
    #             'abr': bairro['abr'],
    #             'mai': bairro['mai'],
    #             'jun': bairro['jun'],
    #             'jul': bairro['jul'],
    #             'ago': bairro['ago'],
    #             'set': bairro['set'],
    #             'out': bairro['out'],
    #             'nov': bairro['nov'],
    #             'dez': bairro['dez'],
    #         },
    #         'total': bairro['total']
    #     }
    #     data.append(bairro_data)

    # query = '''
    # SELECT
    #     bairro,
    #     COUNT(*) AS total,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 1) AS `jan`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 2) AS `fev`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 3) AS `mar`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 4) AS `abr`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 5) AS `mai`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 6) AS `jun`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 7) AS `jul`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 8) AS `ago`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 9) AS `et`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 10) AS `out`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 11) AS `nov`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 12) AS `dez`
    # FROM
    #     iluminacao_ordemdeservico
    # WHERE
    #     bairro IS NOT NULL
    #     AND status IN ('0', '1', '2')
    #     AND YEAR(dt_solicitacao) = 2023
    # GROUP BY
    #     bairro
    # '''

    # with connection.cursor() as cursor:
    #     cursor.execute(query)
    #     rows = cursor.fetchall()

    # data = []
    # for row in rows:
    #     bairro = row[0]
    #     total = row[1]
    #     meses = row[2:]
    #     os_por_mes = {meses[i]: value for i, value in enumerate(meses)}
    #     data.append({'bairro': bairro, 'mes': os_por_mes, 'total': total})

    # query = """
    #     SELECT (SELECT COUNT(*) FROM iluminacao_ordemdeservico WHERE status!='f' and bairro = main.bairro GROUP BY bairro) as total, bairro, (SELECT COUNT(dt_solicitacao) FROM iluminacao_ordemdeservico GROUP BY bairro), 
    #         CASE
    #             WHEN strftime('%m', dt_solicitacao) = '01' THEN 'Janeiro'
    #             WHEN strftime('%m', dt_solicitacao) = '02' THEN 'Fevereiro'
    #             WHEN strftime('%m', dt_solicitacao) = '03' THEN 'Março'
    #             WHEN strftime('%m', dt_solicitacao) = '04' THEN 'Abril'
    #             WHEN strftime('%m', dt_solicitacao) = '05' THEN 'Maio'
    #             WHEN strftime('%m', dt_solicitacao) = '06' THEN 'Junho'
    #             WHEN strftime('%m', dt_solicitacao) = '07' THEN 'Julho'
    #             WHEN strftime('%m', dt_solicitacao) = '08' THEN 'Agosto'
    #             WHEN strftime('%m', dt_solicitacao) = '09' THEN 'Setembro'
    #             WHEN strftime('%m', dt_solicitacao) = '10' THEN 'Outubro'
    #             WHEN strftime('%m', dt_solicitacao) = '11' THEN 'Novembro'
    #             WHEN strftime('%m', dt_solicitacao) = '12' THEN 'Dezembro'
    #             ELSE NULL
    #     END AS nome_mes
    #     FROM iluminacao_ordemdeservico as main
    #     WHERE status != 'f'
    #     GROUP BY bairro
    #     ORDER BY bairro;
    #     """
    query = """
        SELECT (SELECT COUNT(*) FROM iluminacao_ordemdeservico WHERE status!='f' and bairro = main.bairro GROUP BY bairro) as total,bairro, 
        COUNT(dt_solicitacao), 
        MONTHNAME(dt_solicitacao) AS nome_mes 
        FROM iluminacao_ordemdeservico as main 
        WHERE status != 'f' 
        GROUP BY bairro, MONTH(dt_solicitacao) 
        ORDER BY bairro; 
    """
    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()

    # print(results)
    resultados_formatados = []
    for row in results:
        total, bairro, count_dt_solicitacao, nome_mes = row
        resultado_dict = {
            'bairro': bairro,
            'total': total,
            'mes': nome_mes,
            'total_por_mes': count_dt_solicitacao
        }
        resultados_formatados.append(resultado_dict)

    bairros=[]
    for result in resultados_formatados:  
        bairros.append(result['bairro'])
    bairros_ = list(set(bairros))

    resultados_agrupados = []
    for b in bairros_:
        resultados_agrupados.append(
            {'bairro': b}
        )

    for result in resultados_formatados:
        bairro = result['bairro']
        mes = result['mes']
        total_por_mes = result['total_por_mes']
        total = result['total']
        print(bairro, mes, total_por_mes, total)
        for i in resultados_agrupados:
            print(i)
            if i['bairro'] == bairro:
                i[mes]=total_por_mes
                i['total']=total
       
                
    #     else:
            # resultados_agrupados.append({'bairro': bairro, f'{mes}':  total_por_mes, 'total': total})   
    #       for item in resultados_agrupados:
    #           if item.bairro ==  bairro:
    #               item[result.mes]=result.total_por_mes


    # print(resultados_agrupados)
    context = {
        'titulo': apps.get_app_config('iluminacao').verbose_name,   
        'results': resultados_agrupados  
    }
    return render(request, 'iluminacao/painel.html', context)


@login_required
@group_required('os_acesso')
def os_index(request):
    if request.user.is_superuser:
        data=OrdemDeServico.objects.all()
    else:
        data=OrdemDeServico.objects.filter(atendente=Pessoa.objects.get(user=request.user))
    if request.method=='POST':
        valor_da_busca=request.POST['valor_da_busca']
        tipo=request.POST['tipo_da_busca']
        # print(valor_da_busca, tipo)
        if tipo == 'atendente':
            if valor_da_busca=='':
                data=data.filter(atendente=None)
            else:
                data=data.filter(atendente__first_name__icontains=valor_da_busca)
        elif tipo == 'bairro':
            data=data.filter(bairro__icontains=valor_da_busca)
        elif tipo == 'data':
            try:
                valor_da_busca_date = datetime.strptime(valor_da_busca, '%d/%m/%Y').date()            
                data=data.filter(dt_solicitacao__date=valor_da_busca_date.strftime('%Y-%m-%d'))
            except:
                valores=valor_da_busca.split('/')
                # print(valores)
                data=data.filter(dt_solicitacao__year=valores[1], dt_solicitacao__month=valores[0])
        elif tipo == 'protocolo':
            data=data.filter(numero__icontains=valor_da_busca)
        elif tipo == 'rua':
            data=data.filter(logradouro__icontains=valor_da_busca)
        elif tipo == 'status':
            status={'Novo': 0,
            'Aguardando': 1,
            'Em execução': 2,
            'Finalizado': 3}
            data=data.filter(status=status[valor_da_busca.capitalize()])
        elif tipo == 'prioridade':
            prioridades={'Normal': 0,
            'Moderada': 1,
            'Urgente': 2
            }
            data=data.filter(prioridade=prioridades[valor_da_busca.capitalize()])

    paginator = Paginator(data, 30)
    page = request.GET.get('page', 1)
    ordens_de_servico = paginator.get_page(page)
    
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'ordens_de_servico': ordens_de_servico
    }
    return render(request, 'iluminacao/index.html', context)


@login_required
def add_os(request):
    
    form = OS_Form(initial={'tipo': Tipo_OS.objects.get(sigla='IP').id})

    if request.method=='POST':
        form=OS_Form(request.POST)
        if form.is_valid():
            os=form.save(commit=False)
            os.cadastrado_por=Pessoa.objects.get(user=request.user)
            os.save()

            return redirect('iluminacao:os_index')                

    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'form': form,
    }

    return render(request, 'iluminacao/adicionar_os.html', context)

@login_required
@group_required('os_acesso')
def detalhes_os(request, id):
    pessoa = Pessoa.objects.get(user=request.user)
    os = OrdemDeServico.objects.get(id=id)
    form_mensagem = NovaMensagemForm(initial={'os': os.id, 'pessoa': pessoa.id})
    try:
        os_ext=OS_ext.objects.get(os=os)        
    except:
        os_ext = None         
    if request.method=='POST': 
        form_mensagem=NovaMensagemForm(request.POST, request.FILES)
        if form_mensagem.is_valid():
           msg=form_mensagem.save(commit=False)
           msg.os=os
           msg.pessoa=pessoa
           msg.save()
           form_mensagem = NovaMensagemForm(initial={'os': os.id, 'pessoa': pessoa.id})       

    linha_tempo=OS_Linha_Tempo.objects.filter(os=os)
    context={
        'form_mensagem': form_mensagem,
        'linha_tempo': linha_tempo,
        'STATUS': STATUS_CHOICES,
        'PRIORIDADES': PRIORIDADE_CHOICES, 
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'os': os,
        'os_ext': os_ext
    }
    return render(request, 'iluminacao/detalhes_os.html', context)

@login_required
@group_required('os_acesso')
def change_status_os(request, id, opcao):
    os=OrdemDeServico.objects.get(id=id)
    os.status=opcao
    if opcao=='f':
        os.dt_conclusao=datetime.now()
    os.save()
    return redirect('iluminacao:detalhes_os', id=id)

@login_required
@group_required('os_acesso')
def change_prioridade_os(request, id, opcao):
    os=OrdemDeServico.objects.get(id=id)
    os.prioridade=opcao
    os.save()
    return redirect('iluminacao:detalhes_os', id=id)

@login_required
@group_required('os_acesso')
def atender_os(request, id):
    os=OrdemDeServico.objects.get(id=id)
    os.atendente=request.user
    os.save()
    return redirect('iluminacao:detalhes_os', id=id)

@login_required
@group_required('os_acesso')
def funcionarios_listar(request):
    funcionarios=Funcionario_OS.objects.all()
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'funcionarios': funcionarios
    }
    return render(request, 'equipe/funcionarios.html', context)

@login_required
@group_required('os_acesso')
def funcionario_cadastrar(request):
    if request.method=='POST':
        form=Funcionario_Form({'pessoa':request.POST['pessoa'], 'nivel': request.POST['nivel'], 'tipo_os': [1]})
        if form.is_valid():
            form.save()
            funcionario=Funcionario_OS()
            return redirect('iluminacao:funcionarios')
        # else:
            # print(form.errors)
    else:
        form=Funcionario_Form(initial={'tipo_os': Tipo_OS.objects.get(sigla='IP')})
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'form': form
    }
    return render(request, 'equipe/funcionarios_cadastrar.html', context)

@login_required
@group_required('os_acesso')
def funcionario_editar(request, id):
    funcionario=Funcionario_OS.objects.get(id=id)
    form=Funcionario_Form_editar(instance=funcionario)
    if request.method=='POST':
        form=Funcionario_Form_editar(request.POST, instance=funcionario)
        if form.is_valid():
            form.save()
            return redirect('funcionarios')
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'form': form,
        'funcionario': funcionario
    }     
    return render(request, 'equipe/funcionarios_editar.html', context)

@login_required
@group_required('os_acesso')
def funcionario_deletar(request, id):
    funcionario=Funcionario_OS.objects.get(id=id)
    funcionario.delete()

    return redirect('iluminacao:funcionarios')

@login_required
@group_required('os_acesso')
def atribuir_equipe(request, id):
    try:
        instancia=OS_ext.objects.get(os=id)
        form=Equipe_Form(instance=instancia)        
    except Exception as e:
        form=Equipe_Form(initial={'os': id})
        instancia=None
        
    if request.method=='POST':
        if instancia:
            form=Equipe_Form(request.POST, instance=instancia)
        else:
            form=Equipe_Form(request.POST)
        if form.is_valid:
            form.save()
            return redirect('iluminacao:detalhes_os', id)
    context={
            'titulo': apps.get_app_config('iluminacao').verbose_name,   
            'form':form,
        }
    return render(request, 'iluminacao/adicionar_ext.html', context)

@login_required
@group_required('os_acesso')
def pontos_os(request, id):
    instancia=OrdemDeServico.objects.get(id=id)
    form=OS_Form_Ponto(instance=instancia)            
        
    if request.method=='POST':       
        form=OS_Form_Ponto(request.POST, instance=instancia)
        if form.is_valid:
            form.save()
            return redirect('iluminacao:detalhes_os', id)
    context={
            'titulo': apps.get_app_config('iluminacao').verbose_name,   
            'form':form,
        }
    return render(request, 'iluminacao/adicionar_ext.html', context)

from django.db.models import Count

@login_required
@group_required('os_acesso')
def imprimir_os(request, id):
    lista_de_os=[OrdemDeServico.objects.get(id=id)]
    context={
        'lista_de_os': lista_de_os
    }
    return render(request, 'iluminacao/imprimir_os.html', context)

@login_required
@group_required('os_acesso')
def imprimir_varias_os(request, ids):
    ids=ids.split('-')
    ids.pop()
    lista_de_os=[]
    for i in ids:
        lista_de_os.append(OrdemDeServico.objects.get(id=i))
    # print(len(lista_de_os))
    # print(ids)
    context={
        'lista_de_os': lista_de_os
    }
    return render(request, 'iluminacao/imprimir_os.html', context)

@login_required
@group_required('os_acesso')
def graficos(request):
    pontos_por_bairro = OrdemDeServico.objects.values('bairro').annotate(total=Sum('pontos_atendidos')).order_by('-total')[:10]
    os_por_bairro = OrdemDeServico.objects.values('bairro').annotate(total=Count('id')).order_by('-total')[:10]

    finalizados = OrdemDeServico.objects.filter(status='f').count()
    nao_finalizados = OrdemDeServico.objects.exclude(status='f').count()

    # os_por_funcionario = Funcionario_OS.objects.annotate(total=Count('id')).order_by('-total')[:10]
    os_por_funcionario = Funcionario_OS.objects.annotate(total_os=Count('os_ext__os')).order_by('-total_os')
    pontos_por_funcionario = Funcionario_OS.objects.annotate(total_pontos=Sum('os_ext__os__pontos_atendidos')).order_by('-total_pontos')    

    context = {    
        'pontos_por_bairro': pontos_por_bairro,
        'os_por_bairro': os_por_bairro,
        'finalizados': finalizados,
        'nao_finalizados': nao_finalizados,
        'os_por_funcionario': os_por_funcionario,
        'pontos_por_funcionario': pontos_por_funcionario,
        'titulo': apps.get_app_config('iluminacao').verbose_name,
    }
    return render(request, 'iluminacao/graficos.html', context)

@login_required
@group_required('os_acesso')
def graficos_ver_mais(request, tipo, subtipo):
    context = {
        'tipo': tipo,
        'titulo': apps.get_app_config('iluminacao').verbose_name,
    }
    workbook = Workbook()
    if tipo == 'pontos-por-bairro':
        pontos_por_bairro = OrdemDeServico.objects.values('bairro').annotate(total=Sum('pontos_atendidos')).order_by('-total')
        dados = [{'y': item['bairro'], 'total': item['total']} for item in pontos_por_bairro]
        context['dados'] = dados
        context['y'] = 'Bairros'
        context['x'] = 'Pontos'
        
        # Crie uma planilha no workbook
        planilha = workbook.active
        planilha.title = 'Pontos por Bairro'
        # Adicione os cabeçalhos das colunas
        planilha['A1'] = 'Bairros'
        planilha['B1'] = 'Pontos'
        # Preencha os dados
        for index, item in enumerate(dados, start=2):
            planilha.cell(row=index, column=1, value=item['y'])
            planilha.cell(row=index, column=2, value=item['total'])

    elif tipo == 'os-por-bairro':
        os_por_bairro = OrdemDeServico.objects.values('bairro').annotate(total=Count('id')).order_by('-total')
        dados = [{'y': item['bairro'], 'total': item['total']} for item in os_por_bairro]
        context['dados'] = dados
        context['y'] = 'Bairros'
        context['x'] = 'Nº de OS'
        
        planilha = workbook.active
        planilha.title = 'OS por Bairro'
        # Adicione os cabeçalhos das colunas
        planilha['A1'] = 'Bairros'
        planilha['B1'] = 'Nº de OS'
        # Preencha os dados
        for index, item in enumerate(dados, start=2):
            planilha.cell(row=index, column=1, value=item['y'])
            planilha.cell(row=index, column=2, value=item['total'])

    elif tipo == 'pontos-por-funcionario':
        pontos_por_funcionario = Funcionario_OS.objects.annotate(total=Sum('os_ext__os__pontos_atendidos')).order_by('-total')
        dados = [{'y': str(item), 'total': item.total} for item in pontos_por_funcionario]
        context['dados'] = dados
        context['y'] = 'Funcionarios'
        context['x'] = 'Pontos'

         # Crie uma planilha no workbook
        planilha = workbook.active
        planilha.title = 'Pontos por Funcionário'
        # Adicione os cabeçalhos das colunas
        planilha['A1'] = 'Funcionários'
        planilha['B1'] = 'Pontos'
        # Preencha os dados
        for index, item in enumerate(dados, start=2):
            planilha.cell(row=index, column=1, value=item['y'])
            planilha.cell(row=index, column=2, value=item['total'])

    else:
        return redirect('iluminacao:kpi')
    if subtipo=='imprimir':
        return render(request, 'iluminacao/graficos_ver_mais_imprimir.html', context)
    elif subtipo=='download':
        data_atual = datetime.now()
        data_formatada = data_atual.strftime("%d-%m-%y")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={tipo}-{data_formatada}.xlsx'        
        workbook.save(response)
        return response
    return render(request, 'iluminacao/graficos_ver_mais.html', context)

@login_required
@group_required('os_acesso')
def mudadados(request):
    finalizados = OrdemDeServico.objects.filter(status='f')
    count = 0
    for item in finalizados:
        mensagens = OS_Linha_Tempo.objects.filter(os=item)
        for mensagem in mensagens:
            if mensagem.mensagem[0] == '*':
                data = mensagem.mensagem[1:11]
                dt = datetime.strptime(data, "%d/%m/%Y")
                dt = dt.strftime("%Y-%m-%d")
                item.dt_conclusao = dt
                item.save()
                msg = mensagem.mensagem.replace(mensagem.mensagem[0:11], "")
                if msg == "":
                   msg = "Data de conclusão alterada devido a registro retroativo."
                mensagem.mensagem = msg
                mensagem.save()
    return render(request, 'template.html')

def alterar_equipes(request):
    os_finalizadas = OrdemDeServico.objects.filter(status='f')
    for os in os_finalizadas:
        extensoes_dos_OS = OS_ext.objects.filter(os=os)
        for extensao in extensoes_dos_OS:
            if extensao.cod_veiculo != '':
                codigo_do_veiculo = extensao.cod_veiculo.split('/')
                for funcionario in codigo_do_veiculo:
                    funcionario_pessoa = Funcionario_OS.objects.get(funcionario)
                    print(funcionario_pessoa)
                    #funcionarios = Funcionario_OS.objects.filter(nome__startswith=funcionario)
                    #print(funcionarios)

    return render(request, 'template.html')
    

@login_required
@group_required('os_acesso')
def salvar_contagem_os(request):
    # Limpa as tabelas existentes antes de salvar os novos dados
    TotalOSPorSemanaAno.objects.all().delete()
    TotalOSPorMesAno.objects.all().delete()

    # Chama os métodos estáticos para obter e salvar os dados
    OrdemDeServico.total_os_por_semana_ano()
    OrdemDeServico.total_os_por_mes_ano()

    return redirect('iluminacao:contagem_os')

@login_required
@group_required('os_acesso')
def contagem_os(request):
    # Obtém a contagem de OS por semana e ano
    total_os_semana_ano = TotalOSPorSemanaAno.objects.all()

    # Obtém a contagem de OS por mês
    total_os_mes_ano = TotalOSPorMesAno.objects.all()

    context = {
        'total_os_semana_ano': total_os_semana_ano,
        'total_os_mes': total_os_mes_ano
    }

    return render(request, 'iluminacao/contagem_os.html', context)